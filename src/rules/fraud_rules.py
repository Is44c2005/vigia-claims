"""
=============================================================
hackIAthon 2026 — Reto Aseguradora del Sur
Motor de Reglas de Negocio — fraud_rules.py
=============================================================
Responsabilidades:
  1. Cargar y cruzar las 6 tablas del dataset
  2. Aplicar reglas críticas RF-01 a RF-07
  3. Calcular score ponderado 0-100 por señal activada
  4. Generar explicación textual en español por siniestro
  5. Exportar siniestros_analizados.csv listo para el dashboard
=============================================================
"""

import pandas as pd
import numpy as np
import os
from datetime import datetime
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

# ─────────────────────────────────────────────
# CONFIGURACIÓN DE RUTAS
# ─────────────────────────────────────────────

BASE_DIR     = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DATA_DIR     = os.path.join(BASE_DIR, "data", "synthetic")
OUTPUT_DIR   = os.path.join(BASE_DIR, "data", "processed")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ─────────────────────────────────────────────
# PESOS DE SEÑALES (Sección 7 del documento)
# ─────────────────────────────────────────────

PESOS = {
    # Señal                              código    puntos_max
    "borde_vigencia_critico":           ("S01",    8),   # ≤ 10 días
    "borde_vigencia_medio":             ("S02",    4),   # 11-30 días
    "demora_robo_alta":                 ("S03",    8),   # > 48 horas
    "demora_robo_media":                ("S04",    4),   # 24-48 horas
    "frecuencia_asegurado_alta":        ("S05",    8),   # ≥ 3 siniestros
    "frecuencia_asegurado_media":       ("S06",    4),   # 2 siniestros
    "frecuencia_conductor_alta":        ("S07",    8),   # ≥ 3 en vehículo
    "frecuencia_conductor_media":       ("S08",    4),   # 2 en vehículo
    "frecuencia_vehiculo_alta":         ("S09",    6),   # ≥ 3 siniestros
    "frecuencia_vehiculo_media":        ("S09b",   3),   # 2 siniestros
    "solo_rc_repetida":                 ("S10",    6),   # > 2 eventos RC
    "proveedor_lista_restrictiva":      ("S11",   10),   # en lista negra
    "proveedor_observado":              ("S12",    5),   # > 2 casos observados
    "documentos_incompletos":           ("S13",    4),   # falta doc obligatorio
    "dinamica_sospechosa":              ("S14",    6),   # volcadura / múltiple
    "sin_tercero_identificado":         ("S15",    5),   # sin rastro del tercero
    "documentos_inconsistentes":        ("S16",   10),   # alteración confirmada
    "reporte_tardio_alto":              ("S17",    5),   # > 7 días
    "reporte_tardio_medio":             ("S18",    3),   # 4-7 días
    "narrativa_similar_alta":           ("S19",    8),   # > 85% similitud
    "narrativa_similar_media":          ("S20",    4),   # 70-84% similitud
    "monto_cercano_suma_asegurada":     ("S21",    4),   # > 95% suma asegurada
}

SCORE_MAXIMO_TEORICO = 85.0  # suma de máximos posibles para normalizar a 100

# ─────────────────────────────────────────────
# REGLAS CRÍTICAS RF (Sección 8 del documento)
# ─────────────────────────────────────────────

REGLAS_CRITICAS = {
    "RF-01": {
        "descripcion": "Cobertura Pérdida Total por Robo (PTxRB)",
        "semaforo":    "Rojo",
        "puntos_extra": 30,
    },
    "RF-02": {
        "descripcion": "Evidencia de Falsificación o Adulteración Documental",
        "semaforo":    "Rojo",
        "puntos_extra": 30,
    },
    "RF-03": {
        "descripcion": "Asegurado, Beneficiario o Proveedor en Lista Restrictiva",
        "semaforo":    "Rojo",
        "puntos_extra": 30,
    },
    "RF-04": {
        "descripcion": "Dinámica del Accidente Físicamente Imposible",
        "semaforo":    "Rojo",
        "puntos_extra": 25,
    },
    "RF-05": {
        "descripcion": "Siniestro Extremo al Borde de Vigencia (< 48 hrs = 2 días)",
        "semaforo":    "Amarillo",
        "puntos_extra": 15,
    },
    "RF-06": {
        "descripcion": "Demora Atípica en Denuncia de Robo (> 4 días)",
        "semaforo":    "Amarillo",
        "puntos_extra": 15,
    },
    "RF-07": {
        "descripcion": "Narrativa Idéntica (Clonada) detectada",
        "semaforo":    "Amarillo",
        "puntos_extra": 15,
    },
}

# ─────────────────────────────────────────────
# TEXTOS DE EXPLICACIÓN POR SEÑAL
# ─────────────────────────────────────────────

TEXTOS_SEÑAL = {
    "S01": lambda v: f"El siniestro ocurrió apenas {v} día(s) después de contratar la póliza (umbral crítico: ≤10 días).",
    "S02": lambda v: f"El siniestro ocurrió {v} días después de contratar la póliza (rango de alerta: 11-30 días).",
    "S03": lambda v: f"La denuncia por robo se realizó {v} día(s) después del evento (supera las 48 horas permitidas).",
    "S04": lambda v: f"La denuncia por robo se demoró entre 24 y 48 horas ({v} día(s)).",
    "S05": lambda v: f"El asegurado registra {v} siniestros previos en los últimos 18 meses (umbral: ≥3).",
    "S06": lambda v: f"El asegurado registra {v} siniestros previos en los últimos 18 meses.",
    "S07": lambda v: f"El conductor aparece en {v} siniestros con este vehículo en 18 meses (umbral: ≥3).",
    "S08": lambda v: f"El conductor aparece en {v} siniestros con este vehículo en 18 meses.",
    "S09": lambda v: f"El vehículo registra {v} siniestros en los últimos 18 meses (umbral: ≥3).",
    "S09b":lambda v: f"El vehículo registra {v} siniestros en los últimos 18 meses.",
    "S10": lambda v: f"El asegurado presenta {v} siniestros previos con cobertura exclusiva de Responsabilidad Civil.",
    "S11": lambda v: f"⚠️ El proveedor '{v}' se encuentra en la Lista Restrictiva de la aseguradora.",
    "S12": lambda v: f"El proveedor '{v}' tiene un {v}% de sus casos marcados como observados.",
    "S13": lambda v: f"Documentos faltantes u obligatorios no entregados: {v}.",
    "S14": lambda v: f"La dinámica del accidente ({v}) requiere revisión cruzada minuciosa.",
    "S15": lambda v: f"Siniestro sin tercero identificado: {v}.",
    "S16": lambda v: f"⚠️ Inconsistencias documentales detectadas: {v}.",
    "S17": lambda v: f"El siniestro fue reportado {v} días después del evento (supera los 7 días).",
    "S18": lambda v: f"El siniestro fue reportado con demora moderada de {v} días.",
    "S19": lambda v: f"⚠️ Narrativa con {v}% de similitud con otro reclamo — posible clonación.",
    "S20": lambda v: f"Narrativa con {v}% de similitud con otro reclamo — requiere revisión.",
    "S21": lambda v: f"El monto reclamado representa el {v}% de la suma asegurada (umbral: >95%).",
}

# ─────────────────────────────────────────────
# ACCIÓN SUGERIDA POR SEMÁFORO
# ─────────────────────────────────────────────

ACCION_SUGERIDA = {
    "Verde":    "Continuar con el flujo normal de liquidación.",
    "Amarillo": "Escalar a la Unidad Antifraude para revisión documental antes de proceder.",
    "Rojo":     "⚠️ Escalar INMEDIATAMENTE a la Unidad Antifraude para revisión especializada de campo. NO proceder con el pago.",
}

# ─────────────────────────────────────────────
# 1. CARGA Y CRUCE DE TABLAS
# ─────────────────────────────────────────────

def cargar_datos(data_dir: str = DATA_DIR) -> dict:
    """Carga los 6 CSVs y retorna un dict con los DataFrames."""
    print("📂 Cargando tablas del dataset...")
    tablas = {}
    archivos = {
        "siniestros":  "siniestros.csv",
        "polizas":     "polizas.csv",
        "asegurados":  "asegurados.csv",
        "vehiculos":   "vehiculos.csv",
        "proveedores": "proveedores.csv",
        "documentos":  "documentos.csv",
    }
    for nombre, archivo in archivos.items():
        ruta = os.path.join(data_dir, archivo)
        tablas[nombre] = pd.read_csv(ruta, low_memory=False)
        print(f"   ✅ {archivo:<25} {len(tablas[nombre]):>5} registros")
    return tablas


def construir_vista_enriquecida(tablas: dict) -> pd.DataFrame:
    """
    Une todas las tablas en una vista plana por siniestro.
    Calcula columnas derivadas necesarias para las reglas.
    """
    print("\n🔗 Construyendo vista enriquecida de siniestros...")

    sin = tablas["siniestros"].copy()
    pol = tablas["polizas"].copy()
    ase = tablas["asegurados"].copy()
    veh = tablas["vehiculos"].copy()
    prv = tablas["proveedores"].copy()
    doc = tablas["documentos"].copy()

    # ── Join con pólizas ──────────────────────────────────────────
    pol_cols = ["id_poliza", "fecha_inicio", "fecha_fin", "suma_asegurada",
                "prima", "deducible_usd", "canal_venta", "estado_poliza"]
    sin = sin.merge(pol[pol_cols], on="id_poliza", how="left", suffixes=("", "_pol"))

    # ── Join con asegurados ───────────────────────────────────────
    ase_cols = ["id_asegurado", "segmento", "antiguedad_cliente_años",
                "reclamos_ultimos_12m", "mora_actual", "score_cliente_simulado",
                "es_perfil_riesgo"]
    sin = sin.merge(ase[ase_cols], on="id_asegurado", how="left", suffixes=("", "_ase"))

    # ── Join con proveedores ──────────────────────────────────────
    prv_cols = ["id_proveedor", "nombre_proveedor", "tipo",
                "porcentaje_casos_observados", "en_lista_restrictiva",
                "reclamos_asociados"]
    sin = sin.merge(prv[prv_cols], on="id_proveedor", how="left", suffixes=("", "_prv"))

    # ── Documentos: resumen por siniestro ─────────────────────────
    doc_resumen = doc.groupby("id_siniestro").agg(
        total_docs          = ("id_documento",           "count"),
        docs_no_entregados  = ("entregado",              lambda x: (x == "No").sum()),
        docs_ilegibles      = ("legible",                lambda x: (x == "No").sum()),
        inconsistencias     = ("inconsistencia_detectada", lambda x: (x == "Sí").sum()),
        observaciones_doc   = ("observacion",            lambda x: " | ".join(x.dropna().unique()[:3])),
    ).reset_index()
    sin = sin.merge(doc_resumen, on="id_siniestro", how="left")
    sin["total_docs"]         = sin["total_docs"].fillna(0).astype(int)
    sin["docs_no_entregados"] = sin["docs_no_entregados"].fillna(0).astype(int)
    sin["inconsistencias"]    = sin["inconsistencias"].fillna(0).astype(int)
    sin["observaciones_doc"]  = sin["observaciones_doc"].fillna("")

    # ── Columnas derivadas de fechas ──────────────────────────────
    sin["fecha_ocurrencia"] = pd.to_datetime(sin["fecha_ocurrencia"], errors="coerce")
    sin["fecha_reporte"]    = pd.to_datetime(sin["fecha_reporte"],    errors="coerce")
    sin["fecha_inicio"]     = pd.to_datetime(sin["fecha_inicio"],     errors="coerce")
    sin["fecha_fin"]        = pd.to_datetime(sin["fecha_fin"],        errors="coerce")

    sin["dias_desde_inicio_calc"] = (sin["fecha_ocurrencia"] - sin["fecha_inicio"]).dt.days
    sin["dias_desde_fin_calc"]    = (sin["fecha_fin"] - sin["fecha_ocurrencia"]).dt.days
    sin["dias_reporte_calc"]      = (sin["fecha_reporte"] - sin["fecha_ocurrencia"]).dt.days

    # Usar columna calculada si la original está ausente o inconsistente
    sin["dias_inicio_final"] = sin["dias_desde_inicio_poliza"].fillna(sin["dias_desde_inicio_calc"])
    sin["dias_reporte_final"] = sin["dias_entre_ocurrencia_reporte"].fillna(sin["dias_reporte_calc"])

    # ── Ratio monto / suma asegurada ─────────────────────────────
    sin["suma_asegurada"] = pd.to_numeric(sin["suma_asegurada"], errors="coerce").fillna(1)
    sin["monto_reclamado"] = pd.to_numeric(sin["monto_reclamado"], errors="coerce").fillna(0)
    sin["ratio_monto_suma"] = (sin["monto_reclamado"] / sin["suma_asegurada"] * 100).round(1)

    print(f"   → Vista enriquecida: {len(sin)} filas × {len(sin.columns)} columnas")
    return sin

# ─────────────────────────────────────────────
# 2. SIMILITUD DE NARRATIVAS (NLP básico)
# ─────────────────────────────────────────────

STOPWORDS_ES = [
    "de","la","el","en","y","a","los","del","se","las","por","un","para",
    "con","no","una","su","al","es","lo","como","más","pero","sus","le",
    "ya","o","fue","este","ha","si","porque","esta","son","entre","cuando",
    "muy","sin","sobre","también","me","hasta","donde","quien","desde",
    "todo","durante","uno","ni","contra","ese","que","cual","han","fue",
    "ser","tiene","han","era","sido","estar","hay","había","haber",
]

def calcular_similitud_narrativas(df: pd.DataFrame) -> pd.DataFrame:
    """
    Compara narrativas usando TF-IDF (unigramas + bigramas) + similitud coseno.
    Agrupa por ramo+sucursal para mantener comparaciones en contexto relevante.
    """
    print("\nCalculando similitud de narrativas (TF-IDF + coseno)...")

    df = df.copy()
    df["max_similitud_pct"]    = 0.0
    df["id_siniestro_similar"] = ""
    df["num_similares_alto"]   = 0

    descripciones = df["descripcion"].fillna("").tolist()
    ids           = df["id_siniestro"].tolist()
    ramos         = df["ramo"].tolist()
    sucursales    = df["sucursal"].fillna("").tolist()

    grupos: dict = {}
    for i, (r, s) in enumerate(zip(ramos, sucursales)):
        grupos.setdefault(f"{r}_{s}", []).append(i)

    procesados = 0
    for indices in grupos.values():
        if len(indices) < 2:
            continue

        textos = [descripciones[i].lower().strip() for i in indices]
        if sum(1 for t in textos if t) < 2:
            continue

        try:
            vectorizer  = TfidfVectorizer(
                analyzer    = "word",
                ngram_range = (1, 2),
                min_df      = 1,
                stop_words  = STOPWORDS_ES,
            )
            matriz      = vectorizer.fit_transform(textos)
            sim_matrix  = cosine_similarity(matriz)
        except ValueError:
            continue

        for local_idx, global_idx in enumerate(indices):
            fila = sim_matrix[local_idx].copy()
            fila[local_idx] = 0.0          # excluir auto-similitud

            mejor_local = int(np.argmax(fila))
            max_sim     = float(fila[mejor_local]) * 100

            df.at[global_idx, "max_similitud_pct"]    = round(max_sim, 1)
            df.at[global_idx, "id_siniestro_similar"] = ids[indices[mejor_local]] if max_sim > 0 else ""
            df.at[global_idx, "num_similares_alto"]   = int(np.sum(fila >= 0.85))
            procesados += 1

    casos_similares = (df["max_similitud_pct"] >= 70).sum()
    print(f"   Siniestros procesados: {procesados}")
    print(f"   Narrativas similares (>=70%): {casos_similares}")
    return df

# ─────────────────────────────────────────────
# 3. APLICAR SEÑALES Y CALCULAR SCORE
# ─────────────────────────────────────────────

def aplicar_señales(row: pd.Series) -> tuple[float, list, list]:
    """
    Evalúa una fila (siniestro enriquecido) contra las 21 señales.
    Retorna: (score_raw, señales_activadas, detalles_explicacion)
    """
    score_raw  = 0.0
    señales    = []   # lista de códigos activados: ["S01:8", "S11:10", ...]
    detalles   = []   # textos en español para la explicación

    cobertura       = str(row.get("cobertura", "")).strip()
    ramo            = str(row.get("ramo", "")).strip()
    dias_inicio     = row.get("dias_inicio_final", 999)
    dias_reporte    = row.get("dias_reporte_final", 0)
    historial       = row.get("historial_siniestros_asegurado", 0)
    sin_vehiculo    = row.get("siniestros_vehiculo_18m", 0)
    prv_lista       = str(row.get("en_lista_restrictiva", "No")).strip()
    prv_observado   = row.get("porcentaje_casos_observados", 0.0)
    nombre_prv      = str(row.get("nombre_proveedor", "")).strip()
    docs_completos  = str(row.get("documentos_completos", "Sí")).strip()
    docs_inconsist  = str(row.get("documentos_inconsistentes", "No")).strip()
    inconsist_count = int(row.get("inconsistencias", 0))
    obs_doc         = str(row.get("observaciones_doc", "")).strip()
    similitud       = float(row.get("max_similitud_pct", 0))
    ratio_monto     = float(row.get("ratio_monto_suma", 0))
    suma_asegurada  = float(row.get("suma_asegurada", 1))

    try:
        dias_inicio  = float(dias_inicio)
    except (TypeError, ValueError):
        dias_inicio = 999
    try:
        dias_reporte = float(dias_reporte)
    except (TypeError, ValueError):
        dias_reporte = 0
    try:
        historial    = int(historial)
    except (TypeError, ValueError):
        historial = 0
    try:
        sin_vehiculo = int(sin_vehiculo)
    except (TypeError, ValueError):
        sin_vehiculo = 0

    # ── S01 / S02: Borde de vigencia ─────────────────────────────
    if 0 <= dias_inicio <= 10:
        pts = PESOS["borde_vigencia_critico"][1]
        score_raw += pts
        señales.append(f"S01:{pts}")
        detalles.append(TEXTOS_SEÑAL["S01"](int(dias_inicio)))

    elif 11 <= dias_inicio <= 30:
        pts = PESOS["borde_vigencia_medio"][1]
        score_raw += pts
        señales.append(f"S02:{pts}")
        detalles.append(TEXTOS_SEÑAL["S02"](int(dias_inicio)))

    # ── S03 / S04: Demora en denuncia de robo ────────────────────
    if cobertura in ["Robo", "Robo Parcial", "Pérdida Total"]:
        if dias_reporte > 2:
            pts = PESOS["demora_robo_alta"][1]
            score_raw += pts
            señales.append(f"S03:{pts}")
            detalles.append(TEXTOS_SEÑAL["S03"](int(dias_reporte)))
        elif dias_reporte >= 1:
            pts = PESOS["demora_robo_media"][1]
            score_raw += pts
            señales.append(f"S04:{pts}")
            detalles.append(TEXTOS_SEÑAL["S04"](int(dias_reporte)))

    # ── S05 / S06: Frecuencia asegurado ──────────────────────────
    if historial >= 3:
        pts = PESOS["frecuencia_asegurado_alta"][1]
        score_raw += pts
        señales.append(f"S05:{pts}")
        detalles.append(TEXTOS_SEÑAL["S05"](historial))
    elif historial == 2:
        pts = PESOS["frecuencia_asegurado_media"][1]
        score_raw += pts
        señales.append(f"S06:{pts}")
        detalles.append(TEXTOS_SEÑAL["S06"](historial))

    # ── S07 / S08 / S09: Frecuencia vehículo/conductor ───────────
    if ramo == "Vehículos" and sin_vehiculo > 0:
        if sin_vehiculo >= 3:
            pts = PESOS["frecuencia_conductor_alta"][1]
            score_raw += pts
            señales.append(f"S07:{pts}")
            detalles.append(TEXTOS_SEÑAL["S07"](sin_vehiculo))

            pts2 = PESOS["frecuencia_vehiculo_alta"][1]
            score_raw += pts2
            señales.append(f"S09:{pts2}")
            detalles.append(TEXTOS_SEÑAL["S09"](sin_vehiculo))

        elif sin_vehiculo == 2:
            pts = PESOS["frecuencia_conductor_media"][1]
            score_raw += pts
            señales.append(f"S08:{pts}")
            detalles.append(TEXTOS_SEÑAL["S08"](sin_vehiculo))

            pts2 = PESOS["frecuencia_vehiculo_media"][1]
            score_raw += pts2
            señales.append(f"S09b:{pts2}")
            detalles.append(TEXTOS_SEÑAL["S09b"](sin_vehiculo))

    # ── S11 / S12: Proveedor ─────────────────────────────────────
    if prv_lista == "Sí":
        pts = PESOS["proveedor_lista_restrictiva"][1]
        score_raw += pts
        señales.append(f"S11:{pts}")
        detalles.append(TEXTOS_SEÑAL["S11"](nombre_prv))
    elif prv_observado > 0.30:
        pts = PESOS["proveedor_observado"][1]
        score_raw += pts
        señales.append(f"S12:{pts}")
        detalles.append(TEXTOS_SEÑAL["S12"](f"{round(prv_observado*100)}%"))

    # ── S13: Documentos incompletos ───────────────────────────────
    if docs_completos == "No":
        pts = PESOS["documentos_incompletos"][1]
        score_raw += pts
        señales.append(f"S13:{pts}")
        detalle_doc = obs_doc[:80] if obs_doc else "Documentos requeridos no entregados"
        detalles.append(TEXTOS_SEÑAL["S13"](detalle_doc))

    # ── S14: Dinámica sospechosa ──────────────────────────────────
    coberturas_sospechosas = ["Volcadura", "Pérdida Total"]
    palabras_sospechosas   = ["madrugada", "sin testigos", "sin cámaras",
                               "fuga", "no identificado", "huye", "lluvia intensa"]
    descripcion_lower = str(row.get("descripcion", "")).lower()

    if cobertura in coberturas_sospechosas or any(p in descripcion_lower for p in palabras_sospechosas):
        pts = PESOS["dinamica_sospechosa"][1]
        score_raw += pts
        señales.append(f"S14:{pts}")
        detalles.append(TEXTOS_SEÑAL["S14"](cobertura))

    # ── S15: Sin tercero identificado ────────────────────────────
    frases_sin_tercero = [
        "se dio a la fuga", "no existe", "huye el tercer", "sin rastro",
        "no se identificaron", "no existen testigos", "no hay tercero",
        "cámaras presentaban fallas", "no identificado"
    ]
    if any(f in descripcion_lower for f in frases_sin_tercero):
        pts = PESOS["sin_tercero_identificado"][1]
        score_raw += pts
        señales.append(f"S15:{pts}")
        detalles.append(TEXTOS_SEÑAL["S15"]("No se identificó tercero involucrado ni evidencia física"))

    # ── S16: Documentos inconsistentes ───────────────────────────
    if docs_inconsist == "Sí" or inconsist_count > 0:
        pts = PESOS["documentos_inconsistentes"][1]
        score_raw += pts
        señales.append(f"S16:{pts}")
        obs_corta = obs_doc[:100] if obs_doc else "Inconsistencias detectadas en la documentación"
        detalles.append(TEXTOS_SEÑAL["S16"](obs_corta))

    # ── S17 / S18: Reporte tardío ─────────────────────────────────
    if dias_reporte > 7:
        pts = PESOS["reporte_tardio_alto"][1]
        score_raw += pts
        señales.append(f"S17:{pts}")
        detalles.append(TEXTOS_SEÑAL["S17"](int(dias_reporte)))
    elif 4 <= dias_reporte <= 7:
        pts = PESOS["reporte_tardio_medio"][1]
        score_raw += pts
        señales.append(f"S18:{pts}")
        detalles.append(TEXTOS_SEÑAL["S18"](int(dias_reporte)))

    # ── S19 / S20: Similitud narrativa ───────────────────────────
    if similitud >= 85:
        pts = PESOS["narrativa_similar_alta"][1]
        score_raw += pts
        señales.append(f"S19:{pts}")
        detalles.append(TEXTOS_SEÑAL["S19"](round(similitud, 1)))
    elif similitud >= 70:
        pts = PESOS["narrativa_similar_media"][1]
        score_raw += pts
        señales.append(f"S20:{pts}")
        detalles.append(TEXTOS_SEÑAL["S20"](round(similitud, 1)))

    # ── S21: Monto cercano a suma asegurada ───────────────────────
    if ratio_monto >= 95:
        pts = PESOS["monto_cercano_suma_asegurada"][1]
        score_raw += pts
        señales.append(f"S21:{pts}")
        detalles.append(TEXTOS_SEÑAL["S21"](round(ratio_monto, 1)))

    return score_raw, señales, detalles


def evaluar_reglas_criticas(row: pd.Series, señales_activadas: list, score_raw: float) -> tuple[list, float, str | None]:
    """
    Evalúa las reglas críticas RF-01 a RF-07.
    Puede elevar el semáforo o agregar puntos extra.
    Retorna: (reglas_activadas, score_extra, semaforo_forzado)
    """
    reglas   = []
    extra    = 0.0
    semaforo_forzado = None

    cobertura    = str(row.get("cobertura", "")).strip()
    prv_lista    = str(row.get("en_lista_restrictiva", "No")).strip()
    docs_inc     = str(row.get("documentos_inconsistentes", "No")).strip()
    inconsist    = int(row.get("inconsistencias", 0))
    dias_inicio  = row.get("dias_inicio_final", 999)
    dias_reporte = row.get("dias_reporte_final", 0)
    similitud    = float(row.get("max_similitud_pct", 0))
    descripcion  = str(row.get("descripcion", "")).lower()

    try:
        dias_inicio  = float(dias_inicio)
        dias_reporte = float(dias_reporte)
    except (TypeError, ValueError):
        dias_inicio  = 999
        dias_reporte = 0

    # RF-01: Pérdida Total por Robo
    if cobertura in ["Pérdida Total", "Robo"] and "S03" in " ".join(señales_activadas):
        reglas.append("RF-01")
        extra += REGLAS_CRITICAS["RF-01"]["puntos_extra"]
        semaforo_forzado = "Rojo"

    # RF-02: Adulteración documental
    if (docs_inc == "Sí" or inconsist >= 2) and any(
        kw in str(row.get("observaciones_doc", "")).lower()
        for kw in ["anterior", "alterac", "falsif", "no coincide", "ilegible"]
    ):
        reglas.append("RF-02")
        extra += REGLAS_CRITICAS["RF-02"]["puntos_extra"]
        semaforo_forzado = "Rojo"

    # RF-03: Lista Restrictiva
    if prv_lista == "Sí":
        reglas.append("RF-03")
        extra += REGLAS_CRITICAS["RF-03"]["puntos_extra"]
        semaforo_forzado = "Rojo"

    # RF-04: Dinámica físicamente imposible
    frases_imposible = [
        "físicamente imposible", "sin sentido", "ilógico",
        "contradicción", "no coincide con el daño", "sin rastro del tercero"
    ]
    if any(f in descripcion for f in frases_imposible):
        reglas.append("RF-04")
        extra += REGLAS_CRITICAS["RF-04"]["puntos_extra"]
        semaforo_forzado = "Rojo"

    # RF-05: Siniestro en < 2 días de la póliza
    if 0 <= dias_inicio <= 2:
        if "RF-05" not in reglas:
            reglas.append("RF-05")
            extra += REGLAS_CRITICAS["RF-05"]["puntos_extra"]
            if not semaforo_forzado:
                semaforo_forzado = "Amarillo"

    # RF-06: Demora robo > 4 días
    if cobertura in ["Robo", "Robo Parcial"] and dias_reporte > 4:
        reglas.append("RF-06")
        extra += REGLAS_CRITICAS["RF-06"]["puntos_extra"]
        if not semaforo_forzado:
            semaforo_forzado = "Amarillo"

    # RF-07: Narrativa clonada
    if similitud >= 85:
        reglas.append("RF-07")
        extra += REGLAS_CRITICAS["RF-07"]["puntos_extra"]
        if not semaforo_forzado:
            semaforo_forzado = "Amarillo"

    return reglas, extra, semaforo_forzado


def normalizar_score(score_raw: float, maximo: float = SCORE_MAXIMO_TEORICO) -> int:
    return int(min(100, round((score_raw / maximo) * 100)))


def clasificar_semaforo(score: int) -> str:
    if score <= 40:
        return "Verde"
    elif score <= 75:
        return "Amarillo"
    else:
        return "Rojo"


# ─────────────────────────────────────────────
# 4. GENERAR EXPLICACIÓN TEXTUAL
# ─────────────────────────────────────────────

def generar_explicacion(row_result: dict) -> str:
    """
    Construye el párrafo de explicación completo para el analista.
    Este texto es el que aparece en el dashboard y alimenta al agente.
    """
    id_sin      = row_result.get("id_siniestro", "")
    score       = row_result.get("score_motor", 0)
    semaforo    = row_result.get("semaforo_motor", "Verde")
    reglas      = row_result.get("reglas_criticas_activadas", "")
    num_señales = row_result.get("num_señales_motor", 0)
    detalles    = row_result.get("_detalles_texto", [])
    accion      = ACCION_SUGERIDA.get(semaforo, "")
    cobertura   = row_result.get("cobertura", "")
    ramo        = row_result.get("ramo", "")
    monto       = row_result.get("monto_reclamado", 0)
    prv_nombre  = row_result.get("nombre_proveedor", "")

    nivel_texto = {
        "Verde":    "BAJO RIESGO",
        "Amarillo": "RIESGO MEDIO",
        "Rojo":     "ALTO RIESGO",
    }.get(semaforo, "")

    lineas = [
        f"ANÁLISIS DE RIESGO — {id_sin}",
        f"Nivel: {nivel_texto} | Score: {score}/100 | Semáforo: {semaforo}",
        f"Ramo: {ramo} | Cobertura: {cobertura} | Monto reclamado: ${monto:,.2f}",
        f"Proveedor: {prv_nombre}",
        "",
    ]

    if reglas:
        lineas.append(f"⚡ REGLAS CRÍTICAS ACTIVADAS: {reglas}")
        for r in reglas.split(", "):
            r = r.strip()
            if r in REGLAS_CRITICAS:
                lineas.append(f"   • {r}: {REGLAS_CRITICAS[r]['descripcion']}")
        lineas.append("")

    if detalles:
        lineas.append(f"📋 SEÑALES DETECTADAS ({num_señales} señal(es)):")
        for i, d in enumerate(detalles, 1):
            lineas.append(f"   {i}. {d}")
        lineas.append("")

    lineas.append(f"📌 ACCIÓN RECOMENDADA: {accion}")

    if semaforo == "Verde":
        lineas.append("Este siniestro no presenta señales de alerta significativas.")

    return " || ".join([l for l in lineas if l.strip()])


# ─────────────────────────────────────────────
# 5. PIPELINE PRINCIPAL
# ─────────────────────────────────────────────

def ejecutar_motor(data_dir: str = DATA_DIR, output_dir: str = OUTPUT_DIR) -> pd.DataFrame:
    """
    Pipeline completo del motor de reglas.
    Retorna el DataFrame analizado y lo exporta a CSV.
    """
    print("=" * 60)
    print("🚀 MOTOR DE REGLAS — hackIAthon Aseguradora del Sur")
    print("=" * 60)

    # 1. Cargar datos
    tablas = cargar_datos(data_dir)

    # 2. Vista enriquecida
    df = construir_vista_enriquecida(tablas)

    # 3. Similitud de narrativas
    df = calcular_similitud_narrativas(df)

    # 4. Aplicar señales y reglas por siniestro
    print("\n⚙️  Aplicando señales y reglas a 1000 siniestros...")

    resultados = []
    for _, row in df.iterrows():
        score_raw, señales, detalles = aplicar_señales(row)
        reglas_rf, extra_rf, semaforo_forzado = evaluar_reglas_criticas(row, señales, score_raw)

        score_total   = score_raw + extra_rf
        score_final   = normalizar_score(score_total)
        semaforo_final = semaforo_forzado if semaforo_forzado else clasificar_semaforo(score_final)

        # Si hay regla roja forzada, asegurar score ≥ 76
        if semaforo_final == "Rojo":
            score_final = max(score_final, 76)
        elif semaforo_final == "Amarillo":
            score_final = max(41, min(75, score_final))

        res = row.to_dict()
        res.update({
            "score_motor":               score_final,
            "semaforo_motor":            semaforo_final,
            "reglas_criticas_activadas": ", ".join(reglas_rf) if reglas_rf else "",
            "señales_motor":             " | ".join(señales),
            "num_señales_motor":         len(señales),
            "max_similitud_pct":         row.get("max_similitud_pct", 0),
            "id_siniestro_similar":      row.get("id_siniestro_similar", ""),
            "num_similares_alto":        row.get("num_similares_alto", 0),
            "_detalles_texto":           detalles,
        })
        resultados.append(res)

    df_resultado = pd.DataFrame(resultados)

    # 5. Generar explicación textual por siniestro
    print("📝 Generando explicaciones textuales...")
    df_resultado["explicacion_alerta"] = df_resultado.apply(
        lambda r: generar_explicacion(r.to_dict()), axis=1
    )

    # 6. Ranking de prioridad (1 = más urgente)
    df_resultado = df_resultado.sort_values("score_motor", ascending=False)
    df_resultado["prioridad_revision"] = range(1, len(df_resultado) + 1)

    # 7. Seleccionar columnas de salida (limpiar columnas internas)
    columnas_salida = [
        "id_siniestro", "id_poliza", "id_asegurado", "id_proveedor",
        "nombre_proveedor", "tipo", "en_lista_restrictiva",
        "ramo", "cobertura", "fecha_ocurrencia", "fecha_reporte",
        "dias_reporte_final", "dias_inicio_final",
        "monto_reclamado", "monto_estimado", "monto_pagado",
        "suma_asegurada", "ratio_monto_suma",
        "estado", "sucursal",
        "documentos_completos", "documentos_inconsistentes",
        "inconsistencias", "observaciones_doc",
        "historial_siniestros_asegurado", "siniestros_vehiculo_18m",
        "reclamos_ultimos_12m", "score_cliente_simulado",
        "max_similitud_pct", "id_siniestro_similar", "num_similares_alto",
        "score_motor", "semaforo_motor",
        "reglas_criticas_activadas", "señales_motor",
        "num_señales_motor", "prioridad_revision",
        "explicacion_alerta",
        "etiqueta_fraude_simulada",
        "descripcion",
    ]
    columnas_presentes = [c for c in columnas_salida if c in df_resultado.columns]
    df_salida = df_resultado[columnas_presentes].copy()

    # 8. Exportar
    import csv as csv_module
    ruta_salida = os.path.join(output_dir, "siniestros_analizados.csv")
    df_salida.to_csv(ruta_salida, index=False, encoding="utf-8-sig", quoting=csv_module.QUOTE_ALL)

    # También exportar Excel con colores de semáforo
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        ruta_xlsx = os.path.join(output_dir, "siniestros_analizados.xlsx")
        with pd.ExcelWriter(ruta_xlsx, engine="openpyxl") as writer:
            df_salida.to_excel(writer, index=False, sheet_name="Siniestros Analizados")
            ws = writer.sheets["Siniestros Analizados"]
            ws.freeze_panes = "A2"
            # Header azul
            h_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            h_font = Font(color="FFFFFF", bold=True, size=10)
            for cell in ws[1]:
                cell.fill = h_fill
                cell.font = h_font
                cell.alignment = Alignment(horizontal="center")
            # Colores semáforo
            col_sem = list(df_salida.columns).index("semaforo_motor") + 1
            col_sc  = list(df_salida.columns).index("score_motor") + 1
            fills = {"Rojo": ("FFC7CE","9C0006"), "Amarillo": ("FFEB9C","9C5700"), "Verde": ("C6EFCE","276221")}
            for r in range(2, len(df_salida) + 2):
                val = ws.cell(row=r, column=col_sem).value
                if val in fills:
                    bg, fg = fills[val]
                    f = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
                    ft = Font(color=fg, bold=True)
                    for c in [col_sem, col_sc]:
                        ws.cell(row=r, column=c).fill = f
                        ws.cell(row=r, column=c).font = ft
        print(f"  📊 Excel exportado: {ruta_xlsx}")
    except ImportError:
        print("  ⚠️  openpyxl no instalado — solo se generó el CSV")

    # 9. Reporte final
    dist = df_salida["semaforo_motor"].value_counts()
    rojos    = df_salida[df_salida["semaforo_motor"] == "Rojo"]
    amarillos = df_salida[df_salida["semaforo_motor"] == "Amarillo"]
    verdes   = df_salida[df_salida["semaforo_motor"] == "Verde"]

    print("\n" + "=" * 60)
    print("✅ MOTOR DE REGLAS COMPLETADO")
    print("=" * 60)
    print(f"  🟢 Verde:    {dist.get('Verde', 0):4d} siniestros")
    print(f"  🟡 Amarillo: {dist.get('Amarillo', 0):4d} siniestros")
    print(f"  🔴 Rojo:     {dist.get('Rojo', 0):4d} siniestros")
    print(f"\n  📊 Score promedio Rojo:     {rojos['score_motor'].mean():.1f}")
    print(f"  📊 Score promedio Amarillo: {amarillos['score_motor'].mean():.1f}")
    print(f"  📊 Score promedio Verde:    {verdes['score_motor'].mean():.1f}")

    rf_counts = {}
    for rf_list in df_salida["reglas_criticas_activadas"].dropna():
        for rf in str(rf_list).split(", "):
            rf = rf.strip()
            if rf.startswith("RF-"):
                rf_counts[rf] = rf_counts.get(rf, 0) + 1

    if rf_counts:
        print("\n  ⚡ Reglas críticas activadas:")
        for rf, cnt in sorted(rf_counts.items()):
            print(f"     {rf}: {cnt} casos — {REGLAS_CRITICAS[rf]['descripcion']}")

    monto_riesgo = rojos["monto_reclamado"].sum()
    print(f"\n  💰 Monto total en riesgo (Rojos): ${monto_riesgo:,.2f}")
    print(f"\n  📁 Archivo exportado: {ruta_salida}")
    print(f"  📄 Columnas: {len(df_salida.columns)} | Filas: {len(df_salida)}")

    # Top 5 más urgentes
    print("\n  🔴 TOP 5 SINIESTROS MÁS URGENTES:")
    top5 = df_salida.head(5)[["id_siniestro", "score_motor", "cobertura",
                               "reglas_criticas_activadas", "num_señales_motor"]]
    for _, r in top5.iterrows():
        rf = f"[{r['reglas_criticas_activadas']}]" if r["reglas_criticas_activadas"] else ""
        print(f"     {r['id_siniestro']} | Score {r['score_motor']:3d} | "
              f"{r['cobertura']:<20} | {r['num_señales_motor']} señales {rf}")

    return df_salida


# ─────────────────────────────────────────────
# 6. FUNCIÓN UTILITARIA: analizar UN siniestro nuevo
# ─────────────────────────────────────────────

def analizar_siniestro_nuevo(siniestro_dict: dict, data_dir: str = DATA_DIR) -> dict:
    """
    Analiza un siniestro nuevo en tiempo real sin re-procesar todo el dataset.
    Útil para la demo del jurado: 'cargue este siniestro y explíquenos el riesgo'.

    Parámetros:
        siniestro_dict: dict con los campos del siniestro nuevo
        data_dir: directorio con los CSVs de referencia

    Retorna: dict con score, semáforo, reglas activadas y explicación
    """
    prv = pd.read_csv(os.path.join(data_dir, "proveedores.csv"))

    row = pd.Series(siniestro_dict)

    # Enriquecer con datos del proveedor si se especificó id_proveedor
    id_prv = siniestro_dict.get("id_proveedor", "")
    if id_prv:
        prv_row = prv[prv["id_proveedor"] == id_prv]
        if not prv_row.empty:
            row["en_lista_restrictiva"]    = prv_row.iloc[0]["en_lista_restrictiva"]
            row["nombre_proveedor"]        = prv_row.iloc[0]["nombre_proveedor"]
            row["porcentaje_casos_observados"] = prv_row.iloc[0]["porcentaje_casos_observados"]
        else:
            row["en_lista_restrictiva"]    = "No"
            row["nombre_proveedor"]        = id_prv
            row["porcentaje_casos_observados"] = 0.0

    row["max_similitud_pct"] = siniestro_dict.get("max_similitud_pct", 0)
    row["dias_inicio_final"] = siniestro_dict.get("dias_desde_inicio_poliza", 999)
    row["dias_reporte_final"] = siniestro_dict.get("dias_entre_ocurrencia_reporte", 0)
    row["ratio_monto_suma"] = 0
    if siniestro_dict.get("monto_reclamado") and siniestro_dict.get("suma_asegurada"):
        row["ratio_monto_suma"] = (
            float(siniestro_dict["monto_reclamado"]) /
            float(siniestro_dict["suma_asegurada"]) * 100
        )

    score_raw, señales, detalles = aplicar_señales(row)
    reglas_rf, extra_rf, semaforo_forzado = evaluar_reglas_criticas(row, señales, score_raw)

    score_total  = score_raw + extra_rf
    score_final  = normalizar_score(score_total)
    semaforo_final = semaforo_forzado if semaforo_forzado else clasificar_semaforo(score_final)

    if semaforo_final == "Rojo":
        score_final = max(score_final, 76)

    resultado = {
        "id_siniestro":              siniestro_dict.get("id_siniestro", "NUEVO"),
        "ramo":                      siniestro_dict.get("ramo", ""),
        "cobertura":                 siniestro_dict.get("cobertura", ""),
        "monto_reclamado":           siniestro_dict.get("monto_reclamado", 0),
        "nombre_proveedor":          row.get("nombre_proveedor", ""),
        "score_motor":               score_final,
        "semaforo_motor":            semaforo_final,
        "reglas_criticas_activadas": ", ".join(reglas_rf),
        "señales_motor":             " | ".join(señales),
        "num_señales_motor":         len(señales),
        "_detalles_texto":           detalles,
    }
    resultado["explicacion_alerta"] = generar_explicacion(resultado)
    return resultado


# ─────────────────────────────────────────────
# PUNTO DE ENTRADA
# ─────────────────────────────────────────────

if __name__ == "__main__":
    df_analizado = ejecutar_motor()
